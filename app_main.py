# app.py
# Run:
# pip install -U streamlit requests pandas openpyxl torch transformers scipy numpy sentencepiece langdetect
# Optional (only if you upload .xlsb): pip install pyxlsb
# Optional (only if you upload .ods): pip install odfpy

# ---- place at top of app.py (before imports that download models) ----
import os
from pathlib import Path

DEFAULT_CACHE_ROOT = Path(__file__).parent / "ai_cache"
CACHE_ROOT = Path(os.getenv("AI_CACHE_ROOT", str(DEFAULT_CACHE_ROOT)))
CACHE_ROOT.mkdir(parents=True, exist_ok=True)

os.environ.setdefault("HF_HOME", str(CACHE_ROOT / "hf"))
os.environ.setdefault("TRANSFORMERS_CACHE", str(CACHE_ROOT / "hf" / "transformers"))
os.environ.setdefault("HF_DATASETS_CACHE", str(CACHE_ROOT / "hf" / "datasets"))
os.environ.setdefault("HUGGINGFACE_HUB_CACHE", str(CACHE_ROOT / "hf" / "hub"))
os.environ.setdefault("TORCH_HOME", str(CACHE_ROOT / "torch"))
os.environ.setdefault("WDM_CACHE", str(CACHE_ROOT / "wdm"))
os.environ.setdefault("XDG_CACHE_HOME", str(CACHE_ROOT / ".cache"))
os.environ.setdefault("TEMP", str(CACHE_ROOT / "temp"))
os.environ.setdefault("TMP", str(CACHE_ROOT / "temp"))
# ----------------------------------------------------------------------

import io, re, gc, time, random, logging, json, html as _html
from urllib.parse import urlparse
from typing import List, Dict, Tuple, Optional, Any
from dataclasses import dataclass, field

import numpy as np
import pandas as pd
import requests
import streamlit as st
import torch
from langdetect import detect
from scipy.special import softmax
from transformers import (
    AutoTokenizer,
    AutoModelForSequenceClassification,
    pipeline as hf_pipeline,
    AutoModelForSeq2SeqLM,
    M2M100Tokenizer,
)
from openpyxl.utils import get_column_letter

# ---------- environment for stability ----------
os.environ.setdefault("TOKENIZERS_PARALLELISM", "false")
os.environ.setdefault("OMP_NUM_THREADS", "4")
os.environ.setdefault("MKL_NUM_THREADS", "4")
torch.set_grad_enabled(False)

# ------------------ App config ------------------
st.set_page_config(
    page_title="Vivino → Translate → Emotions + Entropy (Gated)",
    page_icon="🍷",
    layout="wide",
)
logging.getLogger("urllib3").setLevel(logging.WARNING)

# ------------------ Session state ------------------
if "stage" not in st.session_state:
    # idle → fetched → analyzed
    st.session_state["stage"] = "idle"

if "df_raw_stage" not in st.session_state:
    st.session_state["df_raw_stage"] = None

if "meta" not in st.session_state:
    st.session_state["meta"] = {}

if "cache" not in st.session_state:
    st.session_state["cache"] = {}

# Separate headers for API (JSON) and webpage (HTML)
HEADERS_API = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
    "Accept": "application/json",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.vivino.com/",
    "Origin": "https://www.vivino.com",
    "X-Requested-With": "XMLHttpRequest",
}
HEADERS_HTML = {
    "User-Agent": HEADERS_API["User-Agent"],
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": HEADERS_API["Accept-Language"],
    "Referer": "https://www.vivino.com/",
}

PAUSE_BETWEEN_REQUESTS = (1.0, 3.0)
RATE_LIMIT_SLEEP = 30

EMOTION_MODEL = "cardiffnlp/twitter-roberta-base-emotion-latest"
TRANS_MODELS = {
    "NLLB-200 distilled (600M, faster)": "facebook/nllb-200-distilled-600M",
    "M2M100-1.2B (more accurate, slower)": "facebook/m2m100_1.2B",
}
NUM_FMT = "0.####################"
EMOTION_COLS = ["anger","anticipation","disgust","fear","joy","love","optimism","pessimism","sadness","surprise","trust"]

# --- standardized emotion settings ---
EMO_MAX_LENGTH = 512
EMO_REVIEW_TOKEN_CAP = 2048
EMO_SENT_BATCH = 32

# --- standardized translation settings ---
TX_MAX_LENGTH = 512
TX_REVIEW_TOKEN_CAP = 2048
TX_SENT_BATCH = 32
TX_MAX_NEW_TOKENS_DEFAULT = 256

# --- stability controls for large datasets ---
SAFE_DISABLE_PREVIEW_N = 2000
SHARD_SIZE = 600
UI_UPDATE_EVERY = 200

# 2-letter → NLLB codes
L2N = {
    "en":"eng_Latn","fr":"fra_Latn","de":"deu_Latn","it":"ita_Latn","es":"spa_Latn","pt":"por_Latn","nl":"nld_Latn",
    "sv":"swe_Latn","no":"nob_Latn","da":"dan_Latn","fi":"fin_Latn","pl":"pol_Latn","cs":"ces_Latn","ro":"ron_Latn",
    "hu":"hun_Latn","el":"ell_Grek","tr":"tur_Latn","ru":"rus_Cyrl","uk":"ukr_Cyrl","bg":"bul_Cyrl","hr":"hrv_Latn",
    "sr":"srp_Cyrl","sk":"slk_Latn","sl":"slv_Latn","he":"heb_Hebr","iw":"heb_Hebr",
}

# =========================
# 🔌 PATCH SYSTEM (kept)
# =========================
PATCH_REGISTRY: Dict[str, List[Tuple[int, str, Any]]] = {}

def register_patch(hook: str, name: Optional[str] = None, order: int = 100):
    def deco(fn):
        PATCH_REGISTRY.setdefault(hook, []).append((order, name or fn.__name__, fn))
        PATCH_REGISTRY[hook].sort(key=lambda t: t[0])
        return fn
    return deco

def run_patches(hook: str, ctx: "Ctx"):
    for _, pname, fn in PATCH_REGISTRY.get(hook, []):
        try:
            fn(ctx)
        except Exception as e:
            st.warning(f"Patch '{pname}' failed at '{hook}': {e}")

@dataclass
class Ctx:
    wine_url: Optional[str] = None
    wine_id: Optional[int] = None
    lang_param: Optional[str] = None
    translator_choice: Optional[str] = None
    tx_batch: int = 32
    tx_max_tokens: int = 256
    live_preview: bool = True
    df_raw: Optional[pd.DataFrame] = None
    df_m1: Optional[pd.DataFrame] = None
    df_m2: Optional[pd.DataFrame] = None
    dom_counts: Optional[pd.Series] = None
    avg_score: Optional[float] = None
    m1_mean_entropy: Optional[float] = None
    m2_mean_entropy: Optional[float] = None
    price_raw: Optional[str] = None
    cache: Dict[str, Any] = field(default_factory=dict)

# ------------------ Helpers ------------------
def _coerce_int(x):
    try:
        return int(str(x).strip())
    except Exception:
        return None

def parse_wine_id_from_url(url: str):
    if not url:
        return None
    try:
        path = urlparse(url).path or ""
    except Exception:
        return None
    m = re.search(r"/w(?:ines)?/(\d+)", path)
    if m:
        return int(m.group(1))
    m = re.search(r"[?&]wine_id=(\d+)", url)
    if m:
        return int(m.group(1))
    return None

def parse_year_from_url(url: str) -> Optional[int]:
    if not url:
        return None
    m = re.search(r"[?&]year=(\d{4})\b", url)
    return int(m.group(1)) if m else None

def move_cols_to_front(df: pd.DataFrame, front: List[str]) -> pd.DataFrame:
    front_present = [c for c in front if c in df.columns]
    rest = [c for c in df.columns if c not in front_present]
    return df[front_present + rest]

def df_to_excel_bytes(df: pd.DataFrame, sheet_name="data", float_cols: List[str] = None, num_fmt: str = NUM_FMT) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        if float_cols is None:
            float_cols = [c for c in df.columns if pd.api.types.is_float_dtype(df[c])]
        for idx, col in enumerate(df.columns, 1):
            if col in float_cols:
                for cell in ws[get_column_letter(idx)][1:]:
                    cell.number_format = num_fmt
    buf.seek(0)
    return buf.read()

def get_json(url, params, attempts=6):
    delay = 5
    for k in range(attempts):
        try:
            r = requests.get(url, headers=HEADERS_API, params=params, timeout=30)
            if r.status_code == 429:
                time.sleep(RATE_LIMIT_SLEEP)
                continue
            r.raise_for_status()
            return r.json() or {}
        except requests.exceptions.RequestException:
            if k == attempts - 1:
                raise
            time.sleep(delay)
            delay = min(90, delay * 2)

# ✅ REPLACES read_table_any_excel: supports CSV + all Excel types
def read_table_any(uploaded_file) -> pd.DataFrame:
    """
    Reads: .csv, .tsv, .txt, .xlsx, .xls, .xlsm, .xlsb, .ods
    """
    name = (getattr(uploaded_file, "name", "") or "").lower()

    # rewind stream
    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    # CSV
    if name.endswith(".csv"):
        try:
            return pd.read_csv(uploaded_file)
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, encoding="latin-1")

    # TSV / TXT
    if name.endswith(".tsv") or name.endswith(".txt"):
        try:
            return pd.read_csv(uploaded_file, sep="\t")
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, sep="\t", encoding="latin-1")

    # Excel variants
    if name.endswith(".xlsb"):
        return pd.read_excel(uploaded_file, engine="pyxlsb")  # pip install pyxlsb
    if name.endswith(".ods"):
        return pd.read_excel(uploaded_file, engine="odf")     # pip install odfpy

    return pd.read_excel(uploaded_file)  # xlsx/xls/xlsm

# ------------------ Entropy ------------------
def entropy_per_row(prob_df: pd.DataFrame, base: int = 2) -> pd.Series:
    p = prob_df.to_numpy(float)
    K = p.shape[1]
    row_sums = p.sum(axis=1, keepdims=True)
    bad = ~np.isfinite(row_sums) | (row_sums <= 0)
    p = np.divide(p, row_sums, out=np.full_like(p, 1.0 / K), where=~bad)
    p = np.clip(p, 1e-300, 1.0)
    H = -(p * (np.log(p) / np.log(base))).sum(axis=1)
    return pd.Series(H, index=prob_df.index)

# ------------------ Fetch reviews ------------------
def fetch_reviews(wine_id: int, language: Optional[str], status_cb, year: Optional[int] = None) -> pd.DataFrame:
    url = f"https://www.vivino.com/api/wines/{wine_id}/reviews"
    rows: List[Dict] = []
    page = 1
    got_any = False

    while True:
        params = {"per_page": 50, "page": page}
        if language:
            params["language"] = language
        if year:
            params["year"] = year

        status_cb(f"Fetching page {page} …")
        try:
            payload = get_json(url, params)
            reviews = payload.get("reviews", [])
            if not reviews:
                break
            got_any = True
            for rev in reviews:
                user = rev.get("user") or {}
                note = (rev.get("note") or "").replace("\n", " ").strip()
                ts = rev.get("created_at") or rev.get("updated_at") or ""

                vint = rev.get("vintage") or {}
                vyear = _coerce_int(vint.get("year") or rev.get("vintage_year") or rev.get("year"))

                rows.append({
                    "review_id": rev.get("id"),  # stable join key
                    "wine_id": wine_id,
                    "vintage_year": vyear,
                    "reviewer": user.get("alias") or user.get("seo_name") or "",
                    "review_date": (ts or "").split("T")[0],
                    "review_ts": ts,
                    "review_score": rev.get("rating") or None,
                    "review_text": note
                })
            page += 1
            time.sleep(random.uniform(*PAUSE_BETWEEN_REQUESTS))
        except Exception as e:
            status_cb(f"Error on page {page}: {e}")
            break

    if not got_any:
        status_cb("No reviews found.")
        return pd.DataFrame(columns=[
            "review_id","wine_id","vintage_year","reviewer","review_date","review_ts","review_score","review_text"
        ])

    df = pd.DataFrame(rows)

    if year is not None:
        df["vintage_year"] = df["vintage_year"].fillna(int(year))

    df["review_dt"] = pd.to_datetime(df["review_ts"], errors="coerce", utc=True)
    if df["review_dt"].isna().all():
        df["review_dt"] = pd.to_datetime(df["review_date"], errors="coerce", utc=True)
    df = (df
          .sort_values("review_dt", ascending=False, kind="stable")
          .drop(columns=["review_dt"])
          .reset_index(drop=True))
    return df

# ------------------ Sentence utils ------------------
_SENT_SPLIT = re.compile(r"(?<=[.!?])\s+")

def _split_sentences(text: str) -> List[str]:
    t = (text or "").strip()
    if not t:
        return []
    return _SENT_SPLIT.split(t)

def _tok_len(tok, s: str, max_length: int) -> int:
    return len(tok(s, add_special_tokens=True, truncation=True, max_length=max_length)["input_ids"])

def _cap_by_token_budget(sents: List[str], tok, cap_tokens: int, max_length: int) -> Tuple[List[str], List[int]]:
    kept, lens, total = [], [], 0
    for s in sents:
        L = _tok_len(tok, s, max_length)
        if L <= 0:
            continue
        if total + L > cap_tokens and kept:
            break
        kept.append(s)
        lens.append(L)
        total += L
    if not kept and sents:
        L = max(1, _tok_len(tok, sents[0], max_length))
        kept, lens = [sents[0]], [min(L, cap_tokens)]
    return kept, lens

# ------------------ Emotions ------------------
@st.cache_resource(show_spinner=False)
def load_emotion_method1():
    tok = AutoTokenizer.from_pretrained(EMOTION_MODEL)
    mdl = AutoModelForSequenceClassification.from_pretrained(EMOTION_MODEL)
    mdl.eval()
    labels = [mdl.config.id2label[i] for i in range(mdl.config.num_labels)]
    idx_map = [labels.index(lbl) for lbl in EMOTION_COLS]
    return tok, mdl, idx_map

@st.cache_resource(show_spinner=False)
def load_emotion_method2():
    clf = hf_pipeline("text-classification", model=EMOTION_MODEL, return_all_scores=True, truncation=True)
    tmp = AutoModelForSequenceClassification.from_pretrained(EMOTION_MODEL)
    labels = [tmp.config.id2label[i] for i in range(tmp.config.num_labels)]
    return clf, labels

def emotions_method1(texts: List[str]) -> pd.DataFrame:
    tok, mdl, idx_map = load_emotion_method1()
    rows = []
    with torch.inference_mode():
        for t in texts:
            sents = _split_sentences(t)
            if not sents:
                rows.append([0.0] * len(EMOTION_COLS))
                continue
            kept, lens = _cap_by_token_budget(sents, tok, EMO_REVIEW_TOKEN_CAP, EMO_MAX_LENGTH)
            vec = np.zeros(len(EMOTION_COLS), dtype=np.float64)
            den = float(sum(lens))
            for i0 in range(0, len(kept), EMO_SENT_BATCH):
                chunk = kept[i0 : i0 + EMO_SENT_BATCH]
                enc = tok(chunk, return_tensors="pt", padding=True, truncation=True, max_length=EMO_MAX_LENGTH)
                logits = mdl(**enc).logits.detach().cpu().numpy()
                probs = softmax(logits, axis=1)[:, idx_map]
                w = np.array(lens[i0 : i0 + len(chunk)], dtype=np.float64)[:, None]
                vec += (probs * w).sum(axis=0)
            rows.append((vec / den).tolist() if den > 0 else vec.tolist())
    df = pd.DataFrame(rows, columns=EMOTION_COLS, dtype=np.float64)
    df.insert(0, "dominant_emotion", df[EMOTION_COLS].idxmax(axis=1))
    return df

def emotions_method2(texts: List[str]) -> pd.DataFrame:
    clf, _ = load_emotion_method2()
    tok, _, _ = load_emotion_method1()
    rows = []
    for t in texts:
        sents = _split_sentences(t)
        if not sents:
            rows.append([0.0] * len(EMOTION_COLS))
            continue
        kept, lens = _cap_by_token_budget(sents, tok, EMO_REVIEW_TOKEN_CAP, EMO_MAX_LENGTH)
        den = float(sum(lens))
        vec = np.zeros(len(EMOTION_COLS), dtype=np.float64)
        for i0 in range(0, len(kept), EMO_SENT_BATCH):
            chunk = kept[i0 : i0 + EMO_SENT_BATCH]
            outs = clf(chunk, padding=True, truncation=True, max_length=EMO_MAX_LENGTH, batch_size=16)
            for j, out in enumerate(outs):
                m = {d["label"]: float(d["score"]) for d in out}
                arr = np.array([m.get(lbl, 0.0) for lbl in EMOTION_COLS], dtype=np.float64)
                vec += arr * float(lens[i0 + j])
        rows.append((vec / den).tolist() if den > 0 else vec.tolist())
    df = pd.DataFrame(rows, columns=EMOTION_COLS, dtype=np.float64)
    df.insert(0, "dominant_emotion", df[EMOTION_COLS].idxmax(axis=1))
    return df

# ------------------ Translators ------------------
@st.cache_resource(show_spinner=False)
def load_translator(model_name: str):
    tok = AutoTokenizer.from_pretrained(model_name)
    mdl = AutoModelForSeq2SeqLM.from_pretrained(model_name)
    device = "cuda" if torch.cuda.is_available() else "cpu"
    mdl = mdl.to(device)
    return tok, mdl, device

def _lang_id(tokenizer, code: str) -> int:
    if hasattr(tokenizer, "lang_code_to_id") and tokenizer.lang_code_to_id:
        return tokenizer.lang_code_to_id.get(code) or tokenizer.convert_tokens_to_ids(code)
    if isinstance(tokenizer, M2M100Tokenizer):
        if code == "en":
            return tokenizer.get_lang_id("en")
        return tokenizer.convert_tokens_to_ids(code)
    return tokenizer.convert_tokens_to_ids(code)

def detect_lang_2letter(text: str) -> str:
    try:
        lg = detect(text) if text and str(text).strip() else "en"
        return (lg or "en")[:2]
    except Exception:
        return "en"

def _translate_segments(tok, mdl, device, segs: List[str], forced_bos_token_id: int, batch: int, max_new: int) -> List[str]:
    out = []
    for i0 in range(0, len(segs), batch):
        chunk = segs[i0 : i0 + batch]
        enc = tok(chunk, return_tensors="pt", padding=True, truncation=True, max_length=TX_MAX_LENGTH)
        enc = {k: v.to(device) for k, v in enc.items()}
        gen = mdl.generate(
            **enc,
            forced_bos_token_id=forced_bos_token_id,
            num_beams=1,
            do_sample=False,
            max_new_tokens=max_new,
        )
        dec = tok.batch_decode(gen, skip_special_tokens=True)
        out.extend(dec)
    return out

def translate_batch(texts: List[str], trans_model_key: str, batch_size: int = TX_SENT_BATCH, max_new_tokens: int = TX_MAX_NEW_TOKENS_DEFAULT) -> Tuple[List[str], List[str]]:
    model_name = TRANS_MODELS[trans_model_key]
    tok, mdl, device = load_translator(model_name)
    is_nllb = "nllb" in model_name
    tgt_code = "eng_Latn" if is_nllb else "en"
    tgt_id = _lang_id(tok, tgt_code)

    langs = [detect_lang_2letter(t) for t in texts]
    english_out = [""] * len(texts)

    if is_nllb:
        src_codes = [L2N.get(lg, "eng_Latn") for lg in langs]
        for src in sorted(set(src_codes)):
            idxs = [i for i, s in enumerate(src_codes) if s == src]
            if src == "eng_Latn":
                for i in idxs:
                    english_out[i] = texts[i] or ""
                continue
            tok.src_lang = src
            seg_texts, owners = [], []
            for i in idxs:
                sents = _split_sentences(texts[i] or "")
                kept, _ = _cap_by_token_budget(sents, tok, TX_REVIEW_TOKEN_CAP, TX_MAX_LENGTH)
                for s in kept:
                    seg_texts.append(s)
                    owners.append(i)
            if not seg_texts:
                for i in idxs:
                    english_out[i] = ""
                continue
            dec = _translate_segments(tok, mdl, device, seg_texts, tgt_id, batch_size, max_new_tokens)
            bins: Dict[int, List[str]] = {}
            for seg_out, owner in zip(dec, owners):
                bins.setdefault(owner, []).append(seg_out)
            for i in idxs:
                english_out[i] = " ".join(bins.get(i, []))
    else:
        tgt_id = _lang_id(tok, "en")
        for lg in sorted(set(langs)):
            idxs = [i for i, l in enumerate(langs) if l == lg]
            if lg == "en":
                for i in idxs:
                    english_out[i] = texts[i] or ""
                continue
            tok.src_lang = lg
            seg_texts, owners = [], []
            for i in idxs:
                sents = _split_sentences(texts[i] or "")
                kept, _ = _cap_by_token_budget(sents, tok, TX_REVIEW_TOKEN_CAP, TX_MAX_LENGTH)
                for s in kept:
                    seg_texts.append(s)
                    owners.append(i)
            if not seg_texts:
                for i in idxs:
                    english_out[i] = ""
                continue
            dec = _translate_segments(tok, mdl, device, seg_texts, tgt_id, batch_size, max_new_tokens)
            bins: Dict[int, List[str]] = {}
            for seg_out, owner in zip(dec, owners):
                bins.setdefault(owner, []).append(seg_out)
            for i in idxs:
                english_out[i] = " ".join(bins.get(i, []))
    return langs, english_out

def translate_batch_stream(texts: List[str], trans_model_key: str, batch_size: int = TX_SENT_BATCH, max_new_tokens: int = TX_MAX_NEW_TOKENS_DEFAULT):
    model_name = TRANS_MODELS[trans_model_key]
    tok, mdl, device = load_translator(model_name)
    is_nllb = "nllb" in model_name
    tgt_code = "eng_Latn" if is_nllb else "en"
    tgt_id = _lang_id(tok, tgt_code)

    langs = [detect_lang_2letter(t) for t in texts]
    for i, t in enumerate(texts):
        lg = langs[i]
        if (is_nllb and L2N.get(lg, "eng_Latn") == "eng_Latn") or (not is_nllb and lg == "en"):
            yield i, "en", (t or "")
            continue
        tok.src_lang = L2N.get(lg, "eng_Latn") if is_nllb else lg
        sents = _split_sentences(t or "")
        kept, _ = _cap_by_token_budget(sents, tok, TX_REVIEW_TOKEN_CAP, TX_MAX_LENGTH)
        if not kept:
            yield i, lg, ""
            continue
        dec = _translate_segments(tok, mdl, device, kept, tgt_id, batch_size, max_new_tokens)
        yield i, lg, " ".join(dec)

# ------------------ Price extraction ------------------
def _strip_to_text(html: str) -> str:
    clean = re.sub(r"(?is)<script.*?>.*?</script>|<style.*?>.*?</style>", " ", html)
    clean = re.sub(r"(?is)<[^>]+>", " ", clean)
    clean = _html.unescape(clean)
    clean = re.sub(r"[\s\u00A0]+", " ", clean).strip()
    return clean

def extract_raw_price_string(html: str) -> Optional[str]:
    text = _strip_to_text(html)
    currency_tokens = r"(?:US\$|CA\$|A\$|AU\$|NZ\$|HK\$|S\$|R\$|CHF|DKK|NOK|SEK|PLN|CZK|HUF|RON|TRY|RUB|MXN|ARS|CLP|COP|PEN|ZAR|AED|SAR|QAR|EGP|IDR|MYR|SGD|THB|VND|PHP|KRW|TWD|ILS|BRL|INR|JPY|CNY|RMB|EUR|GBP|USD|[\$\€\£\₹\¥\₩\₽\₺\₴\₱\₦\₪])"
    amount = r"[0-9]{1,3}(?:[.,\s\u00A0][0-9]{3})*(?:[.,][0-9]{1,2})?"
    patterns = [rf"{currency_tokens}\s*{amount}", rf"{amount}\s*{currency_tokens}"]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            return m.group(0).strip()
    return None

def fetch_price_raw_from_link(wine_url: Optional[str], status_cb) -> Optional[str]:
    if not wine_url:
        return None
    try:
        status_cb("Fetching price from link …")
        r = requests.get(wine_url.strip(), headers=HEADERS_HTML, timeout=30)
        if r.status_code == 429:
            time.sleep(RATE_LIMIT_SLEEP)
            r = requests.get(wine_url.strip(), headers=HEADERS_HTML, timeout=30)
        r.raise_for_status()
        return extract_raw_price_string(r.text or "")
    except Exception as e:
        status_cb(f"Price fetch error: {e}")
        return None

# ------------------ Upload translations (CSV/Excel supported) ------------------

def merge_uploaded_translations(df_raw: pd.DataFrame, uploaded_file) -> pd.DataFrame:
    """
    Merge uploaded translations into df_raw and GUARANTEE a clean column named exactly:
      translated_reviews

    Also prints a quick sanity check in Streamlit so you can see whether the upload actually matched.
    """
    up = read_table_any(uploaded_file)
    up.columns = [str(c).strip() for c in up.columns]

    if "translated_reviews" not in up.columns:
        raise ValueError(
            "Uploaded file must contain a column named exactly 'translated_reviews'. "
            "Rename your translation column to translated_reviews."
        )

    # --- prevent translated_reviews_x / translated_reviews_y collisions ---
    base = df_raw.copy()
    if "translated_reviews" in base.columns:
        base = base.drop(columns=["translated_reviews"])

    # --- prefer stable join on review_id ---
    if "review_id" in up.columns and "review_id" in base.columns:
        # normalize dtype to avoid int vs str mismatch
        base["review_id"] = pd.to_numeric(base["review_id"], errors="coerce").astype("Int64")
        up["review_id"] = pd.to_numeric(up["review_id"], errors="coerce").astype("Int64")

        merged = base.merge(
            up[["review_id", "translated_reviews"]],
            on="review_id",
            how="left"
        )

    # --- fallback join on review_text (fragile) ---
    elif "review_text" in up.columns and "review_text" in base.columns:
        merged = base.merge(
            up[["review_text", "translated_reviews"]],
            on="review_text",
            how="left"
        )

    else:
        raise ValueError(
            "Upload must include either 'review_id' (recommended) or 'review_text' to match rows."
        )

    # --- guarantee expected column name exists ---
    if "translated_reviews" not in merged.columns:
        raise RuntimeError(f"Merge produced no translated_reviews. Columns: {list(merged.columns)}")

    # --- sanity checks (so you don't run emotions on blanks) ---
    non_empty = int((merged["translated_reviews"].fillna("").astype(str).str.strip() != "").sum())
    missing = int(merged["translated_reviews"].isna().sum())

    st.info(f"✅ Merge complete: {non_empty} translated rows found, {missing} missing translations.")

    if non_empty == 0:
        st.error(
            "No translated rows matched. Most common causes:\n"
            "- review_id mismatch (different file or IDs changed)\n"
            "- you uploaded a file without review_id/review_text to match\n"
            "- you edited/reordered rows but are trying to match without IDs\n"
        )

    return merged


# ------------------ Render from cache (prevents “reset” after downloads) ------------------
def render_analyzed_from_cache():
    c = st.session_state["cache"]
    if not c or "df_raw" not in c:
        return

    df_raw = c["df_raw"]
    meta = st.session_state["meta"]

    st.header("Results (cached)")
    st.caption("This view persists even after you click download buttons.")

    st.subheader("Translations available")
    st.dataframe(
        df_raw[[col for col in ["review_id","vintage_year","review_text","language","translated_reviews"] if col in df_raw.columns]],
        use_container_width=True,
        height=360
    )

    st.subheader("Summary")
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    with col1: st.metric("Reviews", f"{c.get('n_reviews', len(df_raw))}")
    with col2: st.metric("Avg score", f"{c.get('avg_score', float('nan')):.17g}")
    with col3: st.metric("Price on page", c.get("price_raw") or "—")
    with col4: st.metric("Mean entropy (M1)", f"{c.get('m1_mean_entropy', float('nan')):.6f}")
    with col5: st.metric("Mean entropy (M2)", f"{c.get('m2_mean_entropy', float('nan')):.6f}")
    with col6: st.metric("Vintage", str(meta.get("vintage_year")) if meta.get("vintage_year") else "All")

    st.subheader("Dominant emotions (Method 2)")
    dom_counts = c.get("dom_counts")
    if dom_counts is not None and not getattr(dom_counts, "empty", True):
        st.bar_chart(dom_counts)

    st.subheader("Download")
    st.download_button(
        "Raw + language + translated_reviews.xlsx",
        data=c["raw_bytes"],
        file_name=c["raw_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_raw_cached",
        use_container_width=True
    )
    st.download_button(
        "Emotions - Method 1.xlsx",
        data=c["m1_bytes"],
        file_name=c["m1_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_m1_cached",
        use_container_width=True
    )
    st.download_button(
        "Emotions - Method 2.xlsx",
        data=c["m2_bytes"],
        file_name=c["m2_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_m2_cached",
        use_container_width=True
    )

# ------------------ Analysis runner (stores cache so page doesn’t “reset”) ------------------
def run_emotions_entropy_and_cache(df_raw: pd.DataFrame, meta: dict, status_cb):
    if "translated_reviews" not in df_raw.columns:
        st.error("translated_reviews column missing. Provide translations first.")
        st.stop()

    raw_bytes = df_to_excel_bytes(df_raw, sheet_name="raw_with_translation")

    em_texts = df_raw["translated_reviews"].fillna("").astype(str).tolist()
    N = len(df_raw)

    m1_parts, m2_parts = [], []
    for s in range(0, N, SHARD_SIZE):
        e = min(N, s + SHARD_SIZE)
        status_cb(f"Emotions on reviews {s+1}–{e} of {N}")
        chunk = em_texts[s:e]
        try:
            m1 = emotions_method1(chunk)
            m2 = emotions_method2(chunk)
        except RuntimeError as err:
            if "CUDA out of memory" in str(err):
                st.error("CUDA OOM during emotions. Reduce SHARD_SIZE.")
                torch.cuda.empty_cache()
                st.stop()
            raise
        m1_parts.append(m1)
        m2_parts.append(m2)
        gc.collect()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    m1_all = pd.concat(m1_parts, ignore_index=True)
    m2_all = pd.concat(m2_parts, ignore_index=True)

    df_m1 = pd.concat([df_raw.reset_index(drop=True), m1_all.reset_index(drop=True)], axis=1)
    df_m2 = pd.concat([df_raw.reset_index(drop=True), m2_all.reset_index(drop=True)], axis=1)

    for dfX in (df_m1, df_m2):
        dfX["SUM"] = dfX[EMOTION_COLS].sum(axis=1).astype(np.float64)
        dfX["Entropy_bits"] = entropy_per_row(dfX[EMOTION_COLS]).astype(np.float64)

    m1_bytes = df_to_excel_bytes(df_m1, sheet_name="method1", float_cols=EMOTION_COLS + ["SUM","Entropy_bits"])
    m2_bytes = df_to_excel_bytes(df_m2, sheet_name="method2", float_cols=EMOTION_COLS + ["SUM","Entropy_bits"])

    avg_score = pd.to_numeric(df_raw["review_score"], errors="coerce").astype(float).mean()
    m1_mean_entropy = float(df_m1["Entropy_bits"].mean())
    m2_mean_entropy = float(df_m2["Entropy_bits"].mean())
    dom_counts = df_m2["dominant_emotion"].value_counts().sort_values(ascending=False)

    price_raw = fetch_price_raw_from_link(meta.get("wine_url"), status_cb)

    st.session_state["cache"] = {
        "df_raw": df_raw,
        "raw_bytes": raw_bytes,
        "m1_bytes": m1_bytes,
        "m2_bytes": m2_bytes,
        "raw_name": f"{meta['wine_id']}_raw_with_translation.xlsx",
        "m1_name": f"{meta['wine_id']}_emotions_method1.xlsx",
        "m2_name": f"{meta['wine_id']}_emotions_method2.xlsx",
        "n_reviews": int(N),
        "avg_score": float(avg_score) if np.isfinite(avg_score) else float("nan"),
        "m1_mean_entropy": float(m1_mean_entropy),
        "m2_mean_entropy": float(m2_mean_entropy),
        "dom_counts": dom_counts,
        "price_raw": price_raw or None,
    }

    st.session_state["stage"] = "analyzed"

# ------------------ UI ------------------
st.title("Vivino → Translate → Emotions + Entropy")
st.caption("Step 1 Fetch → Step 2 Download raw → Step 3 Choose translation → Step 4 Analyze. Results persist after downloads.")

# Sidebar controls
with st.sidebar:
    wine_url = st.text_input("Vivino wine URL (paste full URL)", placeholder="https://www.vivino.com/US/en/.../w/83912")
    detected = parse_wine_id_from_url(wine_url.strip()) if wine_url else None
    default_id = 83912 if not detected else int(detected)
    wine_id = st.number_input("Vivino wine_id", min_value=1, step=1, value=default_id, format="%d")
    if wine_url and detected:
        st.info(f"Detected wine_id: {detected}")
    if wine_url and not detected and wine_url:
        st.warning("Could not detect wine_id; using manual field.")

    language_choice = st.selectbox(
        "Vivino reviews language filter",
        ["All","en","de","fr","it","pt","es","nl","sv","no","da","fi"],
        index=0
    )
    lang_param = None if language_choice == "All" else language_choice

    st.divider()
    fetch_btn = st.button("Step 1: Fetch reviews", use_container_width=True)
    reset_btn = st.button("Reset all", use_container_width=True)

if reset_btn:
    st.session_state["stage"] = "idle"
    st.session_state["df_raw_stage"] = None
    st.session_state["meta"] = {}
    st.session_state["cache"] = {}
    st.rerun()

log = st.empty()
def status(msg: str):
    log.info(msg)


# STEP 1: FETCH
if fetch_btn:
    with st.spinner("Fetching reviews…"):
        vintage_year = parse_year_from_url(wine_url or "")
        df_raw = fetch_reviews(int(wine_id), lang_param, status, year=vintage_year)

        if df_raw.empty:
            st.warning("No reviews returned.")
            st.stop()

        df_raw = move_cols_to_front(df_raw, [
            "review_id","vintage_year","review_date","reviewer","review_score","review_text","review_ts","wine_id"
        ])

        st.session_state["df_raw_stage"] = df_raw
        st.session_state["meta"] = {
            "wine_id": int(wine_id),
            "wine_url": wine_url or None,
            "lang_param": lang_param,
            "vintage_year": vintage_year,
        }
        st.session_state["stage"] = "fetched"
        st.rerun()

# STEP 2+: SHOW + DOWNLOAD + CHOOSE PATH
if st.session_state["stage"] in ("fetched", "analyzed") and st.session_state["df_raw_stage"] is not None:
    df_raw = st.session_state["df_raw_stage"]
    meta = st.session_state["meta"]

    st.header("Step 2: Reviews + ratings + vintage")
    st.dataframe(
        df_raw[[c for c in ["review_id","vintage_year","review_date","reviewer","review_score","review_text"] if c in df_raw.columns]],
        use_container_width=True,
        height=420
    )

    raw_mined_bytes = df_to_excel_bytes(df_raw, sheet_name="raw_reviews")
    st.download_button(
        "Download mined raw (reviews + vintage + ratings)",
        data=raw_mined_bytes,
        file_name=f"{meta.get('wine_id','wine')}_raw_reviews.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_raw_mined",
        use_container_width=True
    )

    st.divider()

    st.header("Step 3: Choose translation path")
    choice = st.radio(
        "How do you want to translate?",
        ["Use local model (inside app)", "Translate outside & upload file (CSV/Excel)"],
        index=0,
        horizontal=True
    )

    if choice == "Use local model (inside app)":
        st.subheader("Local translation settings")
        translator_choice = st.selectbox("Translator", list(TRANS_MODELS.keys()), index=0)
        live_preview = st.checkbox("Show live translation preview", value=True)
        tx_batch = st.slider("Translation sentence batch", 4, 128, TX_SENT_BATCH, step=4)
        tx_max_tokens = st.slider("Translation max new tokens (per sentence)", 64, 512, TX_MAX_NEW_TOKENS_DEFAULT, step=32)

        go_translate = st.button("Step 4: Translate + Analyze (Emotions + Entropy)", use_container_width=True)

        if go_translate:
            with st.spinner("Translating…"):
                texts_src = df_raw["review_text"].fillna("").astype(str).tolist()
                N = len(df_raw)

                if N >= SAFE_DISABLE_PREVIEW_N:
                    live_preview = False
                    st.info(f"Safe mode: live preview disabled for N={N}.")

                langs_all = [None] * N
                english_all = [""] * N

                if live_preview:
                    st.subheader("Live translation")
                    table = st.empty()
                    pbar = st.progress(0.0)
                    done = 0
                    for i, lg, en in translate_batch_stream(
                        texts_src,
                        translator_choice,
                        batch_size=tx_batch,
                        max_new_tokens=tx_max_tokens
                    ):
                        langs_all[i] = lg
                        english_all[i] = en
                        done += 1
                        if done % UI_UPDATE_EVERY == 0 or done == N:
                            pbar.progress(done / max(1, N))
                            live_df = pd.DataFrame({
                                "vintage_year": df_raw["vintage_year"].tolist()[:done] if "vintage_year" in df_raw.columns else [None] * done,
                                "review_text": texts_src[:done],
                                "language": langs_all[:done],
                                "translated_reviews": english_all[:done]
                            })
                            table.dataframe(live_df, use_container_width=True, height=360)
                    del table
                else:
                    langs_chunks, en_chunks = [], []
                    for s in range(0, N, SHARD_SIZE):
                        e = min(N, s + SHARD_SIZE)
                        status(f"Translating reviews {s+1}–{e} of {N}")
                        langs, english = translate_batch(
                            texts_src[s:e],
                            translator_choice,
                            batch_size=tx_batch,
                            max_new_tokens=tx_max_tokens
                        )
                        langs_chunks.append(langs)
                        en_chunks.append(english)
                        gc.collect()
                        if torch.cuda.is_available():
                            torch.cuda.empty_cache()
                    langs_all = [x for chunk in langs_chunks for x in chunk]
                    english_all = [x for chunk in en_chunks for x in chunk]

                df_raw = df_raw.copy()
                df_raw["language"] = langs_all
                df_raw["translated_reviews"] = english_all

                st.session_state["df_raw_stage"] = df_raw

            with st.spinner("Analyzing emotions + entropy…"):
                run_emotions_entropy_and_cache(df_raw, meta, status)
                st.rerun()

    else:
        st.subheader("Upload your translated file (CSV or Excel)")
        st.caption("Required column name: translated_reviews. Recommended: review_id for matching.")

        uploaded = st.file_uploader(
            "Upload translated file",
            type=["csv", "tsv", "txt", "xlsx", "xls", "xlsm", "xlsb", "ods"]
        )

        if uploaded is not None:
            st.info(f"Uploaded: {uploaded.name}")

            # show columns to reduce confusion (optional)
            try:
                tmp = read_table_any(uploaded)
                tmp.columns = [str(c).strip() for c in tmp.columns]
                st.write("Detected columns:", list(tmp.columns))
            except Exception:
                pass

        go_upload = st.button("Step 4: Merge upload + Analyze (Emotions + Entropy)", use_container_width=True)

        if go_upload:
            if uploaded is None:
                st.warning("Please upload a file first.")
                st.stop()

            with st.spinner("Merging translations…"):
                try:
                    df_merged = merge_uploaded_translations(df_raw, uploaded)
                except Exception as e:
                    st.error(f"Upload/merge failed: {e}")
                    st.stop()

                if "language" not in df_merged.columns:
                    df_merged["language"] = df_merged["review_text"].apply(detect_lang_2letter)

                missing = int(pd.isna(df_merged["translated_reviews"]).sum())
                if missing > 0:
                    st.warning(
                        f"{missing} rows did not match a translation. "
                        "If possible, include review_id in your uploaded file and keep it identical to the mined raw download."
                    )

                st.session_state["df_raw_stage"] = df_merged

            with st.spinner("Analyzing emotions + entropy…"):
                run_emotions_entropy_and_cache(df_merged, meta, status)
                st.rerun()

else:
    if st.session_state["stage"] == "idle":
        st.info("Step 1: Paste a Vivino URL in the sidebar and click **Fetch reviews**.")

# ------------------ FINAL RESULTS ------------------
if st.session_state["stage"] == "analyzed" and st.session_state["cache"].get("raw_bytes"):
    st.divider()
    render_analyzed_from_cache()
